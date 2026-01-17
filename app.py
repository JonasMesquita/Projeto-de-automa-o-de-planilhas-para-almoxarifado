import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import font
from openpyxl import Workbook, load_workbook
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import os

# ---------------- CONFIG ----------------
root = tk.Tk()
root.title("Controle de Estoque")
root.geometry("1200x700")

style = ttk.Style()
style.theme_use("clam")

default_font = font.nametofont("TkDefaultFont")
default_font.configure(size=10)
root.option_add("*Font", default_font)

root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=3)
root.rowconfigure(0, weight=1)

arquivo_excel = None

# ---------------- FUNÇÕES ----------------
def criar_planilha():
    global arquivo_excel
    caminho = filedialog.asksaveasfilename(defaultextension=".xlsx")
    if not caminho:
        return

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Entradas"
    ws1.append(["Data", "Codigo", "Produto", "Quantidade"])

    ws2 = wb.create_sheet("Saidas")
    ws2.append(["Data", "Codigo", "Produto", "Quantidade"])

    wb.save(caminho)
    arquivo_excel = caminho
    atualizar_tudo()

def selecionar_planilha():
    global arquivo_excel
    arquivo_excel = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if arquivo_excel:
        atualizar_tudo()

def registrar(tipo):
    if not arquivo_excel:
        messagebox.showwarning("Aviso", "Selecione ou crie uma planilha")
        return

    codigo = entry_codigo.get()
    produto = entry_produto.get()
    try:
        qtd = int(entry_qtd.get())
    except:
        messagebox.showerror("Erro", "Quantidade inválida")
        return

    wb = load_workbook(arquivo_excel)
    ws = wb["Entradas"] if tipo == "Entrada" else wb["Saidas"]

    ws.append([
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        codigo,
        produto,
        qtd
    ])

    wb.save(arquivo_excel)
    atualizar_tudo()

def atualizar_tudo():
    tree_hist.delete(*tree_hist.get_children())
    tree_resumo.delete(*tree_resumo.get_children())

    wb = load_workbook(arquivo_excel)
    resumo = {}

    # Percorre Entradas e Saídas
    for nome, sinal in [("Entradas", 1), ("Saidas", -1)]:
        for linha in wb[nome].iter_rows(min_row=2, values_only=True):
            data, cod, prod, qtd = linha

            if cod not in resumo:
                resumo[cod] = {
                    "produto": prod,
                    "saldo": 0
                }

            resumo[cod]["saldo"] += qtd * sinal

            # Histórico
            tree_hist.insert(
                "",
                "end",
                values=(data, nome[:-1], cod, prod, qtd)
            )

    # Resumo com alerta
    for cod, info in resumo.items():
        saldo = info["saldo"]

        if saldo <= 10:
            tree_resumo.insert(
                "",
                "end",
                values=(cod, info["produto"], saldo),
                tags=("alerta",)
            )
        else:
            tree_resumo.insert(
                "",
                "end",
                values=(cod, info["produto"], saldo)
            )


def gerar_pdf():
    if not arquivo_excel:
        return

    caminho = filedialog.asksaveasfilename(defaultextension=".pdf")
    if not caminho:
        return

    wb = load_workbook(arquivo_excel)
    doc = SimpleDocTemplate(caminho, pagesize=A4)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("Relatório de Estoque", styles["Title"]))
    elementos.append(Spacer(1, 12))

    tabela = [["Código", "Produto", "Saldo"]]
    resumo = {}

    for nome, sinal in [("Entradas", 1), ("Saidas", -1)]:
        for _, cod, prod, qtd in wb[nome].iter_rows(min_row=2, values_only=True):
            resumo.setdefault(cod, {"produto": prod, "saldo": 0})
            resumo[cod]["saldo"] += qtd * sinal

    for cod, info in resumo.items():
        tabela.append([cod, info["produto"], info["saldo"]])

    elementos.append(Table(tabela, style=[
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey)
    ]))

    doc.build(elementos)
    messagebox.showinfo("PDF", "PDF gerado")

# ---------------- LAYOUT ----------------
main = ttk.Frame(root, padding=10)
main.grid(row=0, column=0, columnspan=2, sticky="nsew")
main.columnconfigure(0, weight=1)
main.columnconfigure(1, weight=3)
main.rowconfigure(1, weight=1)

# ESQUERDA (operação)
form = ttk.LabelFrame(main, text="Movimentação", padding=10)
form.grid(row=0, column=0, sticky="ew")

entry_codigo = ttk.Entry(form)
entry_produto = ttk.Entry(form)
entry_qtd = ttk.Entry(form)

for i, (lbl, ent) in enumerate([
    ("Código", entry_codigo),
    ("Produto", entry_produto),
    ("Quantidade", entry_qtd)
]):
    ttk.Label(form, text=lbl).grid(row=i, column=0, sticky="w")
    ent.grid(row=i, column=1, sticky="ew")

form.columnconfigure(1, weight=1)

btns = ttk.Frame(form)
btns.grid(row=3, column=0, columnspan=2, sticky="ew")
btns.columnconfigure((0,1), weight=1)

ttk.Button(btns, text="Entrada", command=lambda: registrar("Entrada")).grid(row=0, column=0, sticky="ew")
ttk.Button(btns, text="Saída", command=lambda: registrar("Saída")).grid(row=0, column=1, sticky="ew")

# HISTÓRICO (embaixo)
hist = ttk.LabelFrame(main, text="Histórico", padding=5)
hist.grid(row=1, column=0, sticky="nsew")

hist_container = ttk.Frame(hist)
hist_container.pack(expand=True, fill="both")

tree_hist = ttk.Treeview(
    hist_container,
    columns=("Data","Tipo","Código","Produto","Qtd"),
    show="headings"
)
for col in tree_hist["columns"]:
    tree_hist.heading(col, text=col)

scroll_hist = ttk.Scrollbar(
    hist_container,
    orient="vertical",
    command=tree_hist.yview
)
tree_hist.configure(yscrollcommand=scroll_hist.set)

tree_hist.pack(side="left", expand=True, fill="both")
scroll_hist.pack(side="right", fill="y")

# DIREITA (resumo)
right = ttk.LabelFrame(main, text="Resumo Geral", padding=10)
right.grid(row=0, column=1, rowspan=2, sticky="nsew")
right.rowconfigure(0, weight=1)

resumo_container = ttk.Frame(right)
resumo_container.pack(expand=True, fill="both")

tree_resumo = ttk.Treeview(
    resumo_container,
    columns=("Código","Produto","Saldo"),
    show="headings"
)
for col in tree_resumo["columns"]:
    tree_resumo.heading(col, text=col)
tree_resumo.tag_configure(
    "alerta",
    background="#ffcccc",
    foreground="#900000"
)


scroll_resumo = ttk.Scrollbar(
    resumo_container,
    orient="vertical",
    command=tree_resumo.yview
)
tree_resumo.configure(yscrollcommand=scroll_resumo.set)

tree_resumo.pack(side="left", expand=True, fill="both")
scroll_resumo.pack(side="right", fill="y")

btn_right = ttk.Frame(right)
btn_right.pack(fill="x", pady=6)

ttk.Button(
    btn_right,
    text="Selecionar Planilha",
    command=selecionar_planilha
).pack(fill="x", pady=2)

ttk.Button(
    btn_right,
    text="Criar Planilha",
    command=criar_planilha
).pack(fill="x", pady=2)

ttk.Button(
    btn_right,
    text="Gerar PDF",
    command=gerar_pdf
).pack(fill="x", pady=2)


root.mainloop()
